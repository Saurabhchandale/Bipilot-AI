from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReportExporter:
    """Create a multi-sheet Excel report for a Bipilot AI analysis."""

    HEADER_FILL = PatternFill("solid", fgColor="102A43")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(size=16, bold=True, color="102A43")

    def export(
        self,
        output_path: str | Path,
        original_filename: str,
        cleaned_dataframe: pd.DataFrame,
        cleaning_summary: dict[str, Any],
        analysis: dict[str, Any],
        profile: dict[str, Any],
        detection: dict[str, Any],
        ai_insights: dict[str, Any],
        chart_specs: list[dict[str, Any]],
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._overview_sheet(
                writer,
                original_filename,
                cleaning_summary,
                analysis,
                detection,
            )
            self._ai_insights_sheet(writer, ai_insights)
            self._columns_sheet(writer, analysis)
            self._quality_sheet(writer, profile)
            self._outliers_sheet(writer, profile)
            self._correlations_sheet(writer, profile)
            self._chart_plan_sheet(writer, chart_specs)
            self._cleaned_preview_sheet(writer, cleaned_dataframe)

            workbook = writer.book
            for worksheet in workbook.worksheets:
                self._format_sheet(worksheet)

        return output_path

    def _overview_sheet(
        self,
        writer: pd.ExcelWriter,
        original_filename: str,
        cleaning_summary: dict[str, Any],
        analysis: dict[str, Any],
        detection: dict[str, Any],
    ) -> None:
        metadata = analysis.get("metadata", {})
        overview = [
            ["Report", "Bipilot AI Dataset Analysis"],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Source File", original_filename],
            ["Dataset Type", detection.get("detected_type", "Unknown")],
            ["Detection Confidence", f"{detection.get('confidence', 0)}%"],
            ["Rows", metadata.get("row_count", 0)],
            ["Columns", metadata.get("column_count", 0)],
            ["Duplicate Rows Removed", cleaning_summary.get("duplicate_rows_removed", 0)],
            ["Total Missing Values After Cleaning", metadata.get("total_missing_values", 0)],
            ["Recommended KPIs", ", ".join(detection.get("recommended_kpis", []))],
        ]
        pd.DataFrame(overview, columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="Overview", index=False
        )

    def _ai_insights_sheet(self, writer: pd.ExcelWriter, ai_insights: dict[str, Any]) -> None:
        rows = [["Provider", ai_insights.get("provider", "local")], ["Summary", ai_insights.get("summary", "")]]

        for section in ["key_findings", "risks", "recommendations"]:
            for item in ai_insights.get(section, []):
                rows.append([section.replace("_", " ").title(), item])

        provider_note = ai_insights.get("provider_note")
        if provider_note:
            rows.append(["Provider Note", provider_note])

        pd.DataFrame(rows, columns=["Section", "Insight"]).to_excel(
            writer, sheet_name="AI Insights", index=False
        )

    def _columns_sheet(self, writer: pd.ExcelWriter, analysis: dict[str, Any]) -> None:
        rows = []
        for column_type, columns in analysis.get("columns", {}).items():
            for column in columns:
                rows.append([column_type.title(), column])

        if not rows:
            rows.append(["None", "No detected columns"])

        pd.DataFrame(rows, columns=["Column Type", "Column Name"]).to_excel(
            writer, sheet_name="Columns", index=False
        )

    def _quality_sheet(self, writer: pd.ExcelWriter, profile: dict[str, Any]) -> None:
        rows = []
        unique_values = profile.get("unique_values", {})

        for column, null_details in profile.get("null_analysis", {}).items():
            rows.append(
                [
                    column,
                    null_details.get("missing_count", 0),
                    null_details.get("missing_percent", 0),
                    unique_values.get(column, {}).get("unique_count", 0),
                    unique_values.get(column, {}).get("unique_percent", 0),
                ]
            )

        pd.DataFrame(
            rows,
            columns=[
                "Column",
                "Missing Count",
                "Missing Percent",
                "Unique Count",
                "Unique Percent",
            ],
        ).to_excel(writer, sheet_name="Data Quality", index=False)

    def _outliers_sheet(self, writer: pd.ExcelWriter, profile: dict[str, Any]) -> None:
        rows = []
        for column, details in profile.get("outliers", {}).items():
            rows.append(
                [
                    column,
                    details.get("method"),
                    details.get("outlier_count", 0),
                    details.get("outlier_percent", 0),
                    details.get("lower_bound"),
                    details.get("upper_bound"),
                ]
            )

        if not rows:
            rows.append(["No numeric outlier profile available", "", 0, 0, "", ""])

        pd.DataFrame(
            rows,
            columns=[
                "Column",
                "Method",
                "Outlier Count",
                "Outlier Percent",
                "Lower Bound",
                "Upper Bound",
            ],
        ).to_excel(writer, sheet_name="Outliers", index=False)

    def _correlations_sheet(self, writer: pd.ExcelWriter, profile: dict[str, Any]) -> None:
        strong_pairs = profile.get("correlations", {}).get("strong_pairs", [])
        rows = [
            [
                " & ".join(pair.get("columns", [])),
                pair.get("correlation"),
                pair.get("strength"),
            ]
            for pair in strong_pairs
        ]

        if not rows:
            rows.append(["No strong correlations detected", "", ""])

        pd.DataFrame(rows, columns=["Columns", "Correlation", "Strength"]).to_excel(
            writer, sheet_name="Correlations", index=False
        )

    def _chart_plan_sheet(self, writer: pd.ExcelWriter, chart_specs: list[dict[str, Any]]) -> None:
        rows = [
            [
                spec.get("title"),
                spec.get("chart_type"),
                spec.get("x") or spec.get("names") or "",
                spec.get("y") or "",
                spec.get("description", ""),
            ]
            for spec in chart_specs
        ]

        if not rows:
            rows.append(["No chart recommendations", "", "", "", ""])

        pd.DataFrame(
            rows,
            columns=["Title", "Chart Type", "X / Category", "Y / Metric", "Description"],
        ).to_excel(writer, sheet_name="Chart Plan", index=False)

    def _cleaned_preview_sheet(
        self, writer: pd.ExcelWriter, cleaned_dataframe: pd.DataFrame
    ) -> None:
        cleaned_dataframe.head(500).to_excel(writer, sheet_name="Cleaned Preview", index=False)

    def _format_sheet(self, worksheet) -> None:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 60))
            worksheet.column_dimensions[column_letter].width = max(14, max_length + 2)

